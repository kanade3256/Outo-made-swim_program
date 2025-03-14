import os
import openpyxl
import get_ID  # get_ID.py をインポート


# 指定したシートのみを含む新しいExcelファイルを作成し、データを書き込む
def write_to_excel(input_filename, output_filename, sheet_name, cells, names):
    wb = openpyxl.load_workbook(input_filename)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート名 '{sheet_name}' が見つかりません。")

    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)  # 初期シート削除
    ws_original = wb[sheet_name]
    ws_new = new_wb.create_sheet(title=sheet_name)

    # セルの値をコピー
    for row in ws_original.iter_rows():
        for cell in row:
            ws_new[cell.coordinate] = cell.value

    # データを書き込む
    for idx, cell in enumerate(sum(cells, [])):  # cells を平坦化
        if idx < len(names):
            ws_new[cell] = names[idx]

    # 出力先ディレクトリを作成（存在しない場合）
    os.makedirs("data_file", exist_ok=True)

    # ファイルを data_file に保存
    output_path = os.path.join("data_file", output_filename)
    new_wb.save(output_path)
    print(f"Excel ファイルを保存しました: {output_path}")


# 書き込むセルのリスト（距離ごとに異なる）
cell_config = {
    50: ["B", "I", "P", "W", "AD", "AK", "AR", "AY", "BF", "BM"],
    100: ["B", "J", "Q", "Y", "AF", "AN", "AT", "BH", "BO"],
    200: ["B", "L", "V", "AF", "AP"],
    400: ["B", "P", "AD"],
}

# イベントリスト（距離と泳法の組み合わせ）
events = [
    (stroke, distance)
    for stroke in ["im", "fly", "ba", "br", "fr"]
    for distance in [50, 100, 200, 400]
]

# CSVファイルのパス
csv_file = "data_file/test2.csv"

# 各イベントに対して処理を実行
for stroke, distance in events:
    prefixes = cell_config[distance]
    cells = [
        [f"{prefix}{7 + i * 10 + offset}" for offset in [0, -1, 1, -2, 2, -3]]
        for prefix in prefixes
        for i in range(6)
    ]  # 6セット (7~57)

    ids = get_ID.get_player_id(csv_file, (stroke, distance), category="mixed")

    if not ids:  # IDリストが空ならスキップ
        print(f"イベント {stroke}{distance} のIDデータが見つかりません。")
        continue

    # 出力ファイル名を data_file に保存
    output_filename = f"{distance}{stroke}_id.xlsx"

    # シート名を距離に応じて指定
    sheet_name = f"{distance}m"

    write_to_excel(
        "data_file/テンプレート.xlsx", output_filename, sheet_name, cells, ids
    )

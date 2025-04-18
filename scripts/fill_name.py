# fill_name.py
# 役割: Excelの選手IDを基に、選手の名前・フリガナ・学校名・学年を補完する。
# 変数:
#   - player_ids: Excelから取得した選手IDのリスト
#   - id_data_map: IDと対応する選手情報のマッピング
#   - output_dir: 更新されたExcelを保存するディレクトリ

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import logging
import traceback
from module.send_message import send_slack_message

RESULT_DATA_FILE = os.getenv("RESULT_DATA_FILE")
INPUT_DATA_FILE = os.getenv("INPUT_DATA_FILE")
DIRECTORY_PATH = os.getenv("DIRECTORY_PATH", "test/")
MERGED_CSV_DATA_FILE = os.path.join(DIRECTORY_PATH, os.getenv("MERGED_CSV_DATA_FILE"))

def get_player_data_by_id(player_ids, csv_path):
    """
    IDリストに基づき、選手の名前・フリガナ・学校名・学年を取得する。

    引数:
        - player_ids: 選手IDのリスト
        - csv_path: 選手データが格納されたCSVファイルのパス

    戻り値:
        - 選手IDをキーとした辞書 {ID: (氏名, ﾌﾘｶﾞﾅ, 学校名, 学年)}
    """
    df = pd.read_csv(csv_path)

    try:
        df["ID"] = df["ID"].astype(str)
        player_ids = [str(id_) for id_ in player_ids]
    except KeyError:
        raise ValueError("CSV に 'ID' 列が存在しません。")

    id_to_data = {
        row["ID"]: (row["氏名"], row["ﾌﾘｶﾞﾅ"], row["学校名"], row["学年"])
        for _, row in df.iterrows()
    }

    return {id_: id_to_data.get(id_, (None, None, None, None)) for id_ in player_ids}

def update_excel_with_player_data(excel_path, csv_path, target_cells, result_data_file):
    try:
        if not os.path.exists(excel_path):
            logging.error(f"pathが存在しません: {excel_path}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excelファイルが存在しません: {excel_path}")
            return
        if not os.path.exists(csv_path):
            logging.error(f"CSVファイルが見つかりません: {csv_path}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"CSVファイルが見つかりません: {csv_path}")
            return
        wb = load_workbook(excel_path)
        ws = wb.active
        player_ids = [
            str(ws[cell].value) if ws[cell].value is not None else None
            for cell in target_cells
        ]
        id_data_map = get_player_data_by_id(player_ids, csv_path)
        for cell in target_cells:
            col_letter, row_number = coordinate_from_string(cell)
            col_index = column_index_from_string(col_letter)
            value = ws[cell].value
            if value is not None and value in id_data_map:
                name, hurigana, school, grade = id_data_map[value]
                if name is not None:
                    ws[cell].value = name
                if hurigana is not None:
                    ws[f"{get_column_letter(col_index + 1)}{row_number}"].value = hurigana
                if school is not None:
                    ws[f"{get_column_letter(col_index + 2)}{row_number}"].value = school
                if grade is not None:
                    ws[f"{get_column_letter(col_index + 3)}{row_number}"].value = grade
        os.makedirs(result_data_file, exist_ok=True)
        output_path = os.path.join(result_data_file, os.path.basename(excel_path))
        wb.save(output_path)
        logging.info(f"Excelファイルを更新・保存しました: {output_path}")
    except Exception as e:
        logging.error(f"Excel更新処理中にエラー: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excel更新処理中にエラー: {e}\n{traceback.format_exc()}")

def main(result_data_file=None, merged_csv_data_file=None):
    """
    メイン処理:
    1. CSVファイルから選手情報を取得
    2. 競技別にExcelシートを更新
    3. 更新完了のメッセージを出力
    """
    from dotenv import load_dotenv
    load_dotenv()
    result_data_file = result_data_file or os.getenv("RESULT_DATA_FILE")
    directory_path = os.getenv("DIRECTORY_PATH", "test/")
    merged_csv_data_file = merged_csv_data_file or os.path.join(directory_path, os.getenv("MERGED_CSV_DATA_FILE"))
    try:
        cell_config = {
            50: ["B", "I", "P", "W", "AD", "AK", "AR", "AY", "BF", "BM"],
            100: ["B", "J", "Q", "Y", "AF", "AN", "AT", "BH", "BO"],
            200: ["B", "L", "V", "AF", "AP"],
            400: ["B", "P", "AD"],
        }
        events = [
            (stroke, distance)
            for stroke in ["fly", "ba", "br", "fr", "im"]
            for distance in [50, 100, 200, 400]
        ]
        for stroke, distance in events:
            if distance not in cell_config:
                continue
            prefixes = cell_config[distance]
            target_cells_list = [
                f"{prefix}{7 + i * 10 + offset}"
                for prefix in prefixes
                for i in range(6)
                for offset in [0, -1, 1, -2, 2, -3]
            ]
            excel_file = os.path.join(result_data_file, f"{distance}{stroke}_id.xlsx")
            try:
                update_excel_with_player_data(excel_file, merged_csv_data_file, target_cells_list, result_data_file)
                logging.info(f"{stroke}{distance} のExcelファイルの更新が完了しました！")
            except FileNotFoundError as e:
                logging.error(f"ファイルが見つかりません: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルが見つかりません: {e}")
            except Exception as e:
                logging.error(f"{stroke}{distance} の処理中に予期しないエラーが発生しました: {e}", exc_info=True)
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"{stroke}{distance} の処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
    except Exception as e:
        logging.error(f"メイン処理中に予期しないエラーが発生しました: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"メイン処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")

if __name__ == "__main__":
    main()
import os
import logging
import openpyxl
import glob
from typing import List
from scripts.get_ID import get_player_id
from dotenv import load_dotenv
import traceback
from module.send_message import send_slack_message

# ロガーの設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

def partition_names(names: List[str]) -> List[List[str]]:
    """
    名前リスト（早い順）から、最終組を必ず6名にするため、
    リストを反転（シート上で上が遅い、下が速い）し、
    余り（r = len % 6）があればそれを1組目（上側）に、
    残りを6名ずつのグループに分割する。
    """
    # 入力は早い順なので反転して遅い順にする
    names_rev = names[::-1]
    total = len(names_rev)
    r = total % 6
    groups = []
    start = 0
    if r != 0:
        groups.append(names_rev[:r])
        start = r
    while start < total:
        groups.append(names_rev[start : start + 6])
        start += 6
    return groups

def assign_group(group: List[str]) -> List[str]:
    """
    各グループ内で、中心から埋めるように名前を配置する。
    各グループは最終的に6コース（上からコース1～6）に対応するリストを返す。
    割り当てパターン（速い順に並べた場合のコース割り当て）は：
      最速 → コース4  
      2番目 → コース3  
      3番目 → コース5  
      4番目 → コース2  
      5番目 → コース6  
      6番目 → コース1  
    ※グループの人数が6未満の場合は、パターンの先頭分だけ割り当て、残りは空文字列とする。
    """
    pattern = [4, 3, 5, 2, 6, 1]  # 割り当てるコース番号（1～6）
    # グループ内を速い順（降順）にする
    sorted_group = group[::-1]
    # 結果はコース1～6の順（インデックス0～5）
    result = ["" for _ in range(6)]
    for i, name in enumerate(sorted_group):
        if i < len(pattern):
            course = pattern[i]
            result[course - 1] = name
    return result

def write_to_excel(
    input_filename: str, 
    output_filename: str, 
    sheet_name: str, 
    cells: List[List[str]], 
    names: List[str],
    result_data_file: str
) -> bool:
    """
    指定したExcelのシートに選手IDを記入する。

    引数:
        - input_filename: 元となるExcelテンプレートファイル
        - output_filename: 出力するExcelファイル名
        - sheet_name: 書き込むシート名
        - cells: 書き込むセルのリスト（各リストはシート上のコース1～6に対応）
        - names: 書き込む選手IDリスト（早い順）

    戻り値:
        - 書き込みに成功した場合はTrue、失敗した場合はFalse
    """
    try:
        logger.info(f"Excelファイルを読み込み中: {input_filename}")
        try:
            wb = openpyxl.load_workbook(input_filename)
        except FileNotFoundError as e:
            logger.error(f"テンプレートファイルが見つかりません: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"テンプレートファイルが見つかりません: {input_filename} - {e}")
            return False
        except PermissionError as e:
            logger.error(f"テンプレートファイルへのアクセス権限がありません: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"テンプレートファイルへのアクセス権限がありません: {input_filename} - {e}")
            return False
        except Exception as e:
            logger.error(f"ファイルの読み込み中に予期しないエラーが発生しました: {input_filename} - {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルの読み込み中に予期しないエラーが発生しました: {input_filename} - {e}\n{traceback.format_exc()}")
            return False

        if sheet_name not in wb.sheetnames:
            logger.error(f"指定されたシート名が存在しません: '{sheet_name}'")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"指定されたシート名が存在しません: '{sheet_name}'")
            raise ValueError(f"シート名 '{sheet_name}' が見つかりません。")

        # 新しいワークブックを作成し、テンプレートシートの内容をコピーする
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 初期シート削除
        ws_original = wb[sheet_name]
        ws_new = new_wb.create_sheet(title=sheet_name)
        for row in ws_original.iter_rows():
            for cell in row:
                ws_new[cell.coordinate] = cell.value

        print("入力名前リスト（早い順）の最初の20件:", names[:20])
        # グループ分け：シート上は上が遅い人、最終組は6名で速い人
        groups = partition_names(names)
        print("グループごとの人数:", [len(g) for g in groups])
        # 各グループ内を、中心から埋める（各グループ結果はコース1～6の順）に再配置
        assigned_groups = [assign_group(group) for group in groups]
        print("各グループ再配置後（コース1～6順）の内容:")
        for idx, grp in enumerate(assigned_groups, start=1):
            print(f"組 {idx}: {grp}")

        # セル配置情報（cells）とグループごとの名前リストを対応付けて書き込み
        for cell_group, name_group in zip(cells, assigned_groups):
            for cell, name in zip(cell_group, name_group):
                ws_new[cell] = name

        try:
            os.makedirs(result_data_file, exist_ok=True)
            output_path = os.path.join(result_data_file, output_filename)
            new_wb.save(output_path)
            logger.info(f"Excelファイルを保存しました: {output_path}")
            return True
        except PermissionError as e:
            logger.error(f"ファイルへの書き込み権限がありません: {e}")
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルへの書き込み権限がありません: {e}")
            return False
        except Exception as e:
            logger.error(f"Excel書き込み処理中に予期しないエラーが発生しました: {e}", exc_info=True)
            send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excel書き込み処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
            return False
    except Exception as e:
        logger.error(f"Excel書き込み処理全体で予期しないエラーが発生しました: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"Excel書き込み処理全体で予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
        return False

def clean_output_directory(directory: str, pattern: str = "*.xlsx") -> int:
    """
    指定ディレクトリ内の特定パターンに一致するファイルを削除する。
    
    引数:
        - directory: 削除対象のディレクトリ
        - pattern: 削除対象のファイルパターン（デフォルトは *.xlsx）
        
    戻り値:
        - 削除したファイル数
    """
    deleted_count = 0
    try:
        if not os.path.exists(directory):
            logger.warning(f"削除対象ディレクトリが存在しません: {directory}")
            os.makedirs(directory, exist_ok=True)
            return 0

        file_pattern = os.path.join(directory, pattern)
        for file_path in glob.glob(file_pattern):
            try:
                os.remove(file_path)
                deleted_count += 1
                logger.debug(f"ファイルを削除しました: {file_path}")
            except PermissionError as e:
                logger.error(f"ファイルの削除権限がありません: {file_path} - {e}")
                print(f"警告: '{os.path.basename(file_path)}' の削除権限がありません")
            except Exception as e:
                logger.error(f"ファイル削除中にエラーが発生しました: {file_path} - {e}")

        logger.info(f"{directory} から {deleted_count}個のファイルを削除しました")
        return deleted_count
    except Exception as e:
        logger.error(f"ディレクトリのクリーン処理中にエラーが発生しました: {directory} - {e}")
        return deleted_count

def main(result_data_file=None, input_data_file=None, merged_csv_data_file=None, template_file=None) -> None:
    """
    メイン処理:
      1. 競技別に選手ID（名前）を取得
      2. Excelテンプレートの指定シートに選手IDを書き込む
      3. 各イベントごとに処理結果を出力する
    """
    # 引数優先、なければ環境変数
    load_dotenv()
    result_data_file = result_data_file or os.getenv("RESULT_DATA_FILE", "result_output_folder")
    input_data_file = input_data_file or os.getenv("INPUT_DATA_FILE", "input_data_folder")
    merged_csv_data_file = merged_csv_data_file or os.path.join(input_data_file, os.getenv("MERGED_CSV_DATA_FILE", "merged_output.csv"))
    template_file = template_file or os.getenv("TEMPLATE_FILE", "template.xlsx")

    try:
        # 環境変数の検証
        if not result_data_file:
            logger.error("出力先パス(result_data_file)が指定されていません")
            print("エラー: 出力先パス(result_data_file)が指定されていません")
            return
        if not template_file:
            logger.error("テンプレートファイルが指定されていません")
            print("エラー: テンプレートファイルが指定されていません")
            return
        if not os.path.exists(template_file):
            logger.error(f"テンプレートファイルが見つかりません: {template_file}")
            print(f"エラー: テンプレートファイル '{template_file}' が見つかりません")
            return
        if not merged_csv_data_file or not os.path.exists(merged_csv_data_file):
            logger.error(f"マージされたCSVファイルが見つかりません: {merged_csv_data_file}")
            print(f"エラー: マージされたCSVファイル '{merged_csv_data_file}' が見つかりません")
            return

        # 出力フォルダを確認・作成
        if not os.path.exists(result_data_file):
            os.makedirs(result_data_file, exist_ok=True)
            logger.info(f"出力フォルダを作成しました: {result_data_file}")
            print(f"出力フォルダを作成しました: {result_data_file}")

        # 出力フォルダ内の過去データ（Excelファイル）を削除
        deleted_count = clean_output_directory(result_data_file)
        print(f"{result_data_file}内の過去データである.xlsxファイルを{deleted_count}件削除しました。")

        # 各距離に対するセル設定（必要な距離のみ設定）
        cell_config = {
            50: ["B", "I", "P", "W", "AD", "AK", "AR", "AY", "BF", "BM"],
            100: ["B", "J", "Q", "Y", "AF", "AN", "AT", "BH", "BO"],
            200: ["B", "L", "V", "AF", "AP"],
            400: ["B", "P", "AD"],
        }

        # 全種目の組み合わせを生成
        events = [
            (stroke, distance)
            for stroke in ["im", "fly", "ba", "br", "fr"]
            for distance in [50, 100, 200, 400]
            if distance in cell_config  # セル設定がある距離のみ対象
        ]

        successful_events = 0
        failed_events = 0

        # 各種目ごとに選手IDを取得し、Excelに書き込む
        for stroke, distance in events:
            logger.info(f"イベント {stroke}{distance} の処理を開始")

            # セル設定が存在するか確認
            if distance not in cell_config:
                logger.warning(f"距離 {distance}m のセル設定が存在しません")
                continue

            prefixes = cell_config[distance]
            # セルのリストを、各グループ（コース1～6）に対応するよう作成
            # 【修正】行番号の開始位置を 7 から 4 に変更
            cells = [
                [f"{prefix}{4 + i * 10 + j}" for j in range(6)]
                for prefix in prefixes
                for i in range(6)
            ]

            # 選手ID（ここでは名前リスト）を取得する
            ids = get_player_id(merged_csv_data_file, (stroke, distance), category="mixed")
            if not ids:
                logger.warning(f"イベント {stroke}{distance} のIDデータが見つかりません")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} のIDデータが見つかりません。")
                failed_events += 1
                continue

            output_filename = f"{distance}{stroke}_id.xlsx"
            sheet_name = f"{distance}m"
            try:
                success = write_to_excel(template_file, output_filename, sheet_name, cells, ids, result_data_file)
                if success:
                    successful_events += 1
                    logger.info(f"イベント {stroke}{distance} の処理が成功しました")
                else:
                    failed_events += 1
                    logger.error(f"イベント {stroke}{distance} の処理が失敗しました")
            except FileNotFoundError as e:
                logger.error(f"ファイルが見つかりません: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"ファイルが見つかりません: {e}")
                failed_events += 1
                continue
            except ValueError as e:
                logger.error(f"イベント {stroke}{distance} の処理中に値エラーが発生しました: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} の処理中に値エラーが発生しました: {e}")
                failed_events += 1
                continue
            except Exception as e:
                logger.error(f"イベント {stroke}{distance} の処理中に予期しないエラーが発生しました: {e}")
                send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"イベント {stroke}{distance} の処理中に予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")
                failed_events += 1
                continue

        # 処理結果のサマリーを出力
        logger.info(f"処理完了: 成功={successful_events}件, 失敗={failed_events}件")
        if successful_events > 0:
            logger.info(f"処理が完了しました。{successful_events}件のイベントを正常に処理しました。")
        if failed_events > 0:
            logger.warning(f"警告: {failed_events}件のイベントで処理に失敗しました。")

    except Exception as e:
        logger.error(f"予期しないエラーが発生しました: {e}", exc_info=True)
        send_slack_message(os.getenv("APP_NAME", "AquaProgrammer"), f"予期しないエラーが発生しました: {e}\n{traceback.format_exc()}")

if __name__ == "__main__":
    main()

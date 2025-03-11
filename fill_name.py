import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)


# IDから名前、ﾌﾘｶﾞﾅ、学校名、学年を取り出す
def get_player_data_by_id(player_ids, csv_path):
    df = pd.read_csv(csv_path)

    try:
        df["ID"] = df["ID"].astype(str)
        player_ids = [str(id_) for id_ in player_ids]
    except KeyError:
        raise ValueError("CSV に 'ID' 列が存在しません。")

    id_to_data = {
        row["ID"]: (row["氏名"], row["ﾌﾘｶﾞﾅ"], row["学校名"], row["学年"])
        for _, row in df.iterrows()
    }

    return {id_: id_to_data.get(id_, (None, None, None, None)) for id_ in player_ids}


# IDを参照して名前、ﾌﾘｶﾞﾅ、学校名、学年を出力する
def update_excel_with_player_data(excel_path, csv_path, target_cells):
    if not os.path.exists(excel_path):
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    player_ids = [
        str(ws[cell].value) if ws[cell].value is not None else None
        for cell in target_cells
    ]
    id_data_map = get_player_data_by_id(player_ids, csv_path)

    for cell in target_cells:
        col_letter, row_number = coordinate_from_string(cell)
        col_index = column_index_from_string(col_letter)

        value = ws[cell].value
        if value is not None and value in id_data_map:
            name, hurigana, school, grade = id_data_map[value]
            if name is not None:
                ws[cell].value = name
            if hurigana is not None:
                ws[f"{get_column_letter(col_index + 1)}{row_number}"].value = hurigana
            if school is not None:
                ws[f"{get_column_letter(col_index + 2)}{row_number}"].value = school
            if grade is not None:
                ws[f"{get_column_letter(col_index + 3)}{row_number}"].value = grade

    output_dir = "data_file"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, os.path.basename(excel_path))
    wb.save(output_path)


def main():
    csv_file_path = "data_file/test2.csv"
    cell_config = {
        50: ["B", "I", "Q", "X", "AF", "AM", "AU", "BB", "BJ", "BQ"],
        100: ["C", "J", "R", "Y", "AG", "AN", "AV", "BC", "BK", "BR"],
        200: ["D", "K", "S", "Z", "AH", "AO", "AW", "BD", "BL", "BS"],
        400: ["E", "L", "T", "AA", "AI", "AP", "AX", "BE", "BM", "BT"],
    }

    events = [
        (stroke, distance)
        for stroke in ["fly", "ba", "br", "fr", "im"]
        for distance in [50, 100, 200, 400]
    ]

    for stroke, distance in events:
        if distance not in cell_config:
            continue

        prefixes = cell_config[distance]
        target_cells_list = [
            f"{prefix}{7 + i * 10 + offset}"
            for prefix in prefixes
            for i in range(6)
            for offset in [0, -1, 1, -2, 2, -3]
        ]

        excel_file = os.path.join("data_file", f"{distance}{stroke}_id.xlsx")
        update_excel_with_player_data(excel_file, csv_file_path, target_cells_list)
        print(f"{stroke}{distance} のExcelファイルの更新が完了しました！")


if __name__ == "__main__":
    main()
